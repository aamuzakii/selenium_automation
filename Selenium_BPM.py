from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
import os

PATH = "C:\Program Files\chromedriver.exe"
driver = webdriver.Chrome(PATH)

driver.get("http://id.bpm.vivo.xyz/PortalNew/TaskView/Index")

change_lang = driver.find_element_by_partial_link_text('中文')
change_lang.click()
driver.find_elements_by_partial_link_text('Eng')[0].click()

ID = driver.find_element_by_id('txtUserName')
PWD = driver.find_element_by_id('txtPassword')

ID.send_keys("20003134")
PWD.send_keys("123456")
ID.send_keys(Keys.RETURN)
time.sleep(10)

iframe = driver.find_elements_by_tag_name('iframe')[0]
driver.switch_to_frame(iframe)
p = driver.current_window_handle


links = driver.find_elements_by_partial_link_text('Sales Order')
for link in links:
    link.click()

time.sleep(17)
chwd = driver.window_handles

doc_nums = []
excel_names = []


def erase_space(): 
    str_to_list = excel_name.split(' ')

    new_list = []
    def erase_space(variable): 
        if variable != '': 
            return True
        else: 
            return False

    filtered = filter(erase_space, str_to_list) 

    for s in filtered: 
        new_list.append(s)

    str1 = ""      
    for ele in new_list:  
        str1 += ele  
    return str1     



def my_func():
	driver.find_element_by_class_name('more-text').click()
	doc_num = driver.find_elements_by_class_name('profile-span')[7].text
	excel_name = driver.find_element_by_partial_link_text('Order').text  
	excel_name = erase_space(excel_name)
	doc_nums.append(doc_num)
	excel_names.append(excel_name)

for w in chwd:
		if(w!=p):
			driver.switch_to.window(w)
			my_func()


	
df = pd.DataFrame(list(zip(doc_nums,excel_names)))
df = df.rename(columns={0: "BPM Number", 1: "app"})


wb = Workbook()
ws = wb.active
for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)
for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'
wb.save(r'\\172.23.102.26\Users\User\Downloads\BPMtoExcel.xlsx')
os.startfile(r'\\172.23.102.26\Users\User\Downloads\BPMtoExcel.xlsx')



